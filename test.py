# -*- coding:utf-8  -*-
# @Time     : 2022/8/4 22:38
# @Author   : BGLB
# @Software : PyCharm
import json
import os
import time
import traceback
from contextlib import closing
from functools import reduce

import requests
from openpyxl import load_workbook

from core import TaoGongChang

header = {
    'authority': 'tgc.tmall.com',
    'method': 'POST',
    'scheme': 'https',
    'accept': 'application/json, text/plain, */*',
    'accept-language': 'zh-CN,zh;q=0.9',
    'bx-v': '2.2.2',
    'cache-control': 'no-cache',
    'content-type': 'application/json;charset=UTF-8',
    'cookie': 'tfstk=gaZ-Ei2yLsfoNb92YB_m-s8pd960vkerG7y6xXckRSFYQ5roRX2ov9FqiucuE0MKOjwCU7XrVXtQLJi3ETu3MWeL1zD3F8mLJWVarWeFUvGQOWh3rZjgU8oEA1mdjG2yOekCFunBt-wyrHB9kGjgU89oFO2fj9rSw0lZODgSRnOjgvKWAWGWGtHIKQTBVWwfHjkpFYtWPE9jBYHSADNQhtHEpQsZiBM2vyoMaQFMWjxBAoH-PZ2SGHM4DY3-18ZJAH9Ze4h_Fj1grskjlJuLYp-KN8aaTANWNsl3S-qSkDQMTXU3Fuq-Yp8YH2Eu82EPCClTP-EKq7C98fauhmGzP6YSNuyiDygAFgwx2bFaV7IOG5gTqyuL4OOZ6zwYv0qRKUkzl7ox3kf2VXUbRu0rYQ1qtP20Vqn147ZgXkkesfHHytBv8euS38eS5UIYG-frHfX5se8EJfkxstBv8euS3xhGeu8e82hN.; XSRF-TOKEN=e64f22a0-0a1b-4bc3-bb88-dded40bd7514; isg=BIaGbcU9HG0unsgJc2jHe7W113wI58qh2HVXyHCvcqmEcyaN2HcasWwDT6-_W8K5; X-XSRF-TOKEN=18473dc5-1978-4c33-b144-36b791b951f8; SCMBIZTYPE=176000; SCMSESSID=1d9e7855b2d1e1f23b3d29ff80da20a3@HAVANA; xlly_s=1; SCMLOCALE=zh-cn; _nk_=scm326950456; t=3345f427b7b18bf599d4b080bb940c32; locale=zh-cn; _tb_token_=7d1133b7970e3; cookie2=1d9e7855b2d1e1f23b3d29ff80da20a3; lgc=; csg=60f2681b; sgcookie=E100iO48Fo6H34N%2FtmhGrDfXj2f7BuYksZMUnt8xC6zD006mCXzXHFoAWRFMyJRsol77JS2akGjcVIjLv6BPSHn8PaCwQMg61JwYUaLQZqBrepE%3D; cookie17=UUpgTs0ylXKCyUqjpg%3D%3D; sn=; dnk=',
    'origin': 'https://tgc.tmall.com',
    'pragma': 'no-cache',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 ' \
                  'Safari/537.36',
    'x-xsrf-token': '8b1658c2-803d-46da-9e7f-a896dbca2ffa',
}


def get_excel():
    """

    :return:
    """
    url = 'https://tgc.tmall.com/api/v1/orderNew/tradeOrderDownload.htm?spm=a26364.20167850.0.0.59664840JWIIAA&_input_charset=utf-8&taskId=40332923&c2mNavigatorShellPage=2&c2mNavigatorPageOpener=1'
    a = request_download_big_file({'url': url, 'headers': header}, './data.csv')
    print(a)


def request_download_big_file(request_kwargs, local_path, ):
    try:
        if os.path.exists(local_path):
            os.remove(local_path)
        local_path_tmp = local_path+'.tmp'
        param = {'url': '', 'stream': True, 'headers': ''}
        param.update(request_kwargs)
        with closing(requests.get(**param)) as r:
            # r = requests.get(url=file_url, verify=False, stream=True)
            if r.status_code == 200:
                with open(local_path_tmp, "wb") as f:
                    # f.write(r.content)
                    for chunk in r.iter_content(chunk_size=4096):
                        if chunk:
                            f.write(chunk)
            else:
                return False
        os.rename(local_path_tmp, local_path)
    except Exception:
        print(traceback.format_exc())
        return False
    return True


def get_qOsi(mainOrder):
    url = 'https://tgc.tmall.com/ds/api/v1/o/qOsiWithRecord'
    param = {"mainOrderId": '4064869116030936839', "infoKeys": ["buyerNick", "fullName", "mobilephone", "fullAddress"]}
    rep = requests.post(url, json=param, headers=header)
    result = rep.json()

    if result.get('success'):
        return True, result.get('data')
    return False, result.get('errorMessage')


def get_order():
    url = 'https://tgc.tmall.com/api/v1/orderNew/getTradeOrders.htm'
    order_list = []
    page = 1
    pageSize = 10
    param = {
        'pageNo': page,
        'pageSize': pageSize,
        'sourceTradeId': '',
        'status': 'PAID'
    }
    rep = requests.get(url, params=param, headers=header)
    total = rep.json().get('paginator', {}).get('total')
    if total > pageSize:
        while True:
            rep = requests.get(url, params=param, headers=header)
            if rep.json().get('success'):
                order = rep.json().get('data')
                for item in order:
                    print(item)
                    order_list.extend(item.get('detailOrders'))

            if len(order_list) >= total:
                break
            param['pageNo'] = page+1
    else:
        order = rep.json().get('data')
        for item in order:
            order_list.append(item.get('detailOrders'))
    run_function = lambda x, y: x if y in x else x+[y]
    order_list = reduce(run_function, [[], ]+order_list)

    return order_list


def read_excel(path):
    """

    :return:
    """
    res = []
    try:
        wb = load_workbook(path)
        sheet = wb.active
        n = 1
        for row in sheet.iter_rows(values_only=True, min_row=2):
            n += 1
            mainOrder = row[1]
            if mainOrder:
                flag, data = get_qOsi(mainOrder)
                print(f'{mainOrder}: {data}')
                if flag and data:
                    sheet.cell(row=n, column=3, value=data.get('buyerNick'))
                    sheet.cell(row=n, column=12, value=data.get('fullName'))
                    sheet.cell(row=n, column=13, value=data.get('fullAddress'))
                    sheet.cell(row=n, column=14, value=data.get('mobilephone'))
                # time.sleep(.1)
        wb.save('订单数据.xlsx')
    except Exception:
        print(traceback.format_exc())
        return False, traceback.format_exc()
    return True, res

def save_excel(data):
    """
        保存数据
    :param data:
    :return:
    """
    try:
        tmpl_path = os.path.join('tmpl.xlsx')
        wb = load_workbook(tmpl_path)
        sheet = wb.active
        for index, order in enumerate(data):
            order_id = order.get('sourceTradeId')
            index = index+2
            flag, data = get_qOsi(order_id)
            print(f'{order_id}: {data}')
            order.update(data)
            # sheet.cell(row=index, column=1, value=order_id)  # 快递公司
            # sheet.cell(row=index, column=2, value=order_id)  # 快递单号
            sheet.cell(row=index, column=3, value=order_id)  # 订单编号
            # sheet.cell(row=index, column=4, value=order_id)  # 订单来源
            sheet.cell(row=index, column=5, value=order.get('buyerNick'))  # 买家昵称
            sheet.cell(row=index, column=6, value=order.get('fullName'))  # 收货人姓名
            sheet.cell(row=index, column=7, value=order.get('mobilephone'))  # 收货人手机号
            sheet.cell(row=index, column=8, value=order.get('prov'))  # 省
            sheet.cell(row=index, column=9, value=order.get('city'))  # 市
            sheet.cell(row=index, column=10, value=order.get('area'))  # 区/县
            sheet.cell(row=index, column=11, value=order.get('town'))  # 街道地址
            sheet.cell(row=index, column=12, value=order.get('fullAddress'))  # 详细信息
            # sheet.cell(row=index, column=13, value=order.get('mobilephone'))  # 卖家备注
            # sheet.cell(row=index, column=14, value=order.get('mobilephone'))  # 买家留言
            # sheet.cell(row=index, column=15, value=order.get('mobilephone'))  # 实付金额
            sheet.cell(row=index, column=16, value=order.get('auctionTitle'))  # 商品标题
            sheet.cell(row=index, column=17, value=order.get('outerIdSku'))  # 商家编码
            sheet.cell(row=index, column=18, value=order.get('buyAmount'))  # 商品数量
        wb.save('订单数据_{}_{}.xlsx'.format('aa', time.strftime('%Y-%m-%d_%H_%M_%S')))
    except Exception:
        print("生成订单数据失败\n{}".format(traceback.format_exc()))
        return False

def save_data(data):
    """

    :return:
    """
    with open('./data.json', encoding='utf8', mode='r') as f:
        data = json.load(f)
    save_excel(data)


def test():
    all_task = [TaoGongChang, ]
    for task in all_task:
        task.test()


if __name__ == '__main__':
    # a = get_order()
    # save_data(a)
    # print(len(a))
    print(get_qOsi(''))
    # get_order()
